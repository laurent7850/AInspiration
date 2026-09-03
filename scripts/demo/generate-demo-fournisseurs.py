#!/usr/bin/env python3
"""Generate the demonstration workbook for the supplier-invoice consolidation case study.

The real output of that project (`Fournisseurs_Patinoire.xlsx`) carries the restaurant's
purchase prices and the names of its actual suppliers. It must never be screenshotted for
the public showcase. This script rebuilds the same layout — identical columns, fills, fonts,
number formats and column widths — from entirely fictional data, so the capture shows the
product without disclosing anything.

Every supplier name, product code, description and price below is invented.

Usage:
    python scripts/demo/generate-demo-fournisseurs.py [--out <path>]
"""

import argparse
from collections import OrderedDict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ─── Formatting — mirrors extract_invoices.py exactly ────────────────
HEADER_FONT = Font(name='Arial', bold=True, size=11, color='FFFFFF')
HEADER_FILL = PatternFill('solid', fgColor='2F5496')
CATEGORY_FONT = Font(name='Arial', bold=True, size=11, color='2F5496')
CATEGORY_FILL = PatternFill('solid', fgColor='D6E4F0')
DATA_FONT = Font(name='Arial', size=10)
THIN_BORDER = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC'),
)
EURO_FMT = '#,##0.00 €'
TVA_FMT = '0%'
COL_WIDTHS = [16, 50, 12, 15, 8, 15]
HEADERS = ['Code Produit', 'Article / Description', 'Quantité', 'Prix Unitaire', 'TVA', 'Montant']

DEFAULT_OUT = Path(__file__).parent / 'out' / 'Fournisseurs-DEMO.xlsx'

# ─── Fictional data ──────────────────────────────────────────────────
# A product row is (code, description, quantity, unit price, vat rate, amount).
# The first sheet is grouped by category, like the largest supplier of the real file.

CATEGORISED_SHEET = 'Verhaegen Traiteur'

CATEGORISED_PRODUCTS = OrderedDict([
    ('FRAIS', [
        ('F-1042', 'Beurre de baratte doux 250 g', 24, 2.85, 0.06, 68.40),
        ('F-1078', 'Crème épaisse 35 % — 1 L', 12, 4.10, 0.06, 49.20),
        ('F-1155', 'Œufs plein air calibre M — plaque de 30', 8, 6.95, 0.06, 55.60),
        ('F-1203', 'Pâte feuilletée pur beurre 2 kg', 6, 11.40, 0.06, 68.40),
        ('F-1310', 'Lardons fumés 500 g', 10, 4.75, 0.06, 47.50),
    ]),
    ('SEC', [
        ('S-2011', 'Farine T55 — sac 25 kg', 4, 18.90, 0.06, 75.60),
        ('S-2046', 'Riz arborio 5 kg', 6, 13.25, 0.06, 79.50),
        ('S-2088', 'Huile d\'olive vierge extra 5 L', 3, 42.00, 0.06, 126.00),
        ('S-2130', 'Sel de Guérande 1 kg', 12, 2.40, 0.06, 28.80),
        ('S-2177', 'Poivre noir concassé 500 g', 4, 9.80, 0.06, 39.20),
    ]),
    ('CONGELÉ', [
        ('C-3005', 'Frites fraîches surgelées 10 mm — 2,5 kg', 20, 4.35, 0.06, 87.00),
        ('C-3062', 'Petits pois extra-fins 2,5 kg', 8, 5.60, 0.06, 44.80),
        ('C-3118', 'Framboises entières 1 kg', 6, 8.90, 0.06, 53.40),
    ]),
    ('ND', [
        ('', 'Assortiment de mignardises — colis mixte', 2, 34.50, 0.06, 69.00),
    ]),
])

SIMPLE_SHEETS = OrderedDict([
    ('Cave Saint-Roch', [
        ('VIN-0114', 'Côtes du Rhône rouge 2022 — 75 cl', 36, 7.85, 0.21, 282.60),
        ('VIN-0127', 'Chablis 2021 — 75 cl', 24, 14.20, 0.21, 340.80),
        ('VIN-0203', 'Crémant brut — 75 cl', 18, 9.60, 0.21, 172.80),
        ('VIN-0288', 'Sancerre blanc 2022 — 75 cl', 12, 15.75, 0.21, 189.00),
        ('VIN-0341', 'Bordeaux supérieur 2020 — magnum', 6, 21.40, 0.21, 128.40),
    ]),
    ('Marée du Nord', [
        ('POI-0021', 'Filet de cabillaud sans peau — kg', 14.5, 18.90, 0.06, 274.05),
        ('POI-0034', 'Saint-Jacques noix calibre 10/20 — kg', 6.0, 32.50, 0.06, 195.00),
        ('POI-0057', 'Moules de Zélande — sac 5 kg', 8, 14.75, 0.06, 118.00),
        ('POI-0090', 'Crevettes grises décortiquées — kg', 3.5, 46.00, 0.06, 161.00),
    ]),
    ('Primeurs Vanderlinden', [
        ('LEG-0102', 'Pommes de terre bintje — sac 25 kg', 6, 16.40, 0.06, 98.40),
        ('LEG-0118', 'Échalotes grises — kg', 12, 4.20, 0.06, 50.40),
        ('LEG-0145', 'Champignons de Paris bruns — barquette 500 g', 20, 2.35, 0.06, 47.00),
        ('LEG-0170', 'Salade frisée — pièce', 30, 1.15, 0.06, 34.50),
        ('LEG-0192', 'Citrons non traités — kg', 9, 3.05, 0.06, 27.45),
    ]),
    ('Fromagerie Lambrecht', [
        ('FRO-0011', 'Comté affiné 18 mois — kg', 4.2, 24.90, 0.06, 104.58),
        ('FRO-0028', 'Chèvre cendré — pièce 180 g', 16, 3.45, 0.06, 55.20),
        ('FRO-0044', 'Bleu des Causses — kg', 2.5, 21.30, 0.06, 53.25),
    ]),
    ('Distri Boissons', [
        ('BOI-0301', 'Eau plate 50 cl — bac de 24', 15, 9.60, 0.06, 144.00),
        ('BOI-0318', 'Limonade artisanale 33 cl — bac de 24', 8, 16.80, 0.06, 134.40),
        ('BOI-0355', 'Bière blonde locale 33 cl — casier de 24', 12, 22.40, 0.21, 268.80),
        ('BOI-0402', 'Café en grains torréfaction maison — kg', 10, 18.50, 0.06, 185.00),
    ]),
    ('Hygiène Pro', [
        ('HYG-0007', 'Détergent lave-vaisselle professionnel — 20 L', 2, 48.90, 0.21, 97.80),
        ('HYG-0019', 'Essuie-mains pliés — carton de 3 000', 4, 27.60, 0.21, 110.40),
        ('HYG-0033', 'Dégraissant cuisine — 5 L', 3, 19.75, 0.21, 59.25),
    ]),
    # Kept deliberately: in the real file some tabs come from the order sheet only and
    # carry no invoice line yet. Showing one makes the capture truthful about the tool.
    ('Domaine des Trois Ponts', [
        ('', 'Pinot noir 2023 — 75 cl (feuille de commande)', 0, 12.50, 0.21, 0.0),
    ]),
])


def write_header(ws):
    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER
    for col, width in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(col)].width = width


def write_product_row(ws, row, prod):
    code, desc, qty, price, tva, total = prod

    cell = ws.cell(row=row, column=1, value=code if code else '')
    cell.font = DATA_FONT
    cell.border = THIN_BORDER

    cell = ws.cell(row=row, column=2, value=desc)
    cell.font = DATA_FONT
    cell.border = THIN_BORDER

    cell = ws.cell(row=row, column=3, value=qty)
    cell.font = DATA_FONT
    cell.alignment = Alignment(horizontal='center')
    cell.border = THIN_BORDER

    cell = ws.cell(row=row, column=4, value=round(price, 2) if price else 0)
    cell.font = DATA_FONT
    cell.number_format = EURO_FMT
    cell.border = THIN_BORDER

    cell = ws.cell(row=row, column=5, value=tva)
    cell.font = DATA_FONT
    cell.number_format = TVA_FMT
    cell.alignment = Alignment(horizontal='center')
    cell.border = THIN_BORDER

    cell = ws.cell(row=row, column=6, value=round(total, 2) if total else 0)
    cell.font = DATA_FONT
    cell.number_format = EURO_FMT
    cell.border = THIN_BORDER


def write_category_row(ws, row, label):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    cell = ws.cell(row=row, column=1, value=label)
    cell.font = CATEGORY_FONT
    cell.fill = CATEGORY_FILL
    cell.alignment = Alignment(horizontal='left', vertical='center')
    cell.border = THIN_BORDER


def build(out_path):
    wb = Workbook()
    wb.remove(wb.active)

    # Fictional data, but the file should say so where a reader can check without
    # it showing up in a screenshot.
    wb.properties.title = 'Consolidation fournisseurs — jeu de démonstration'
    wb.properties.creator = 'AInspiration'
    wb.properties.description = (
        'Donnees entierement fictives. Genere par scripts/demo/generate-demo-fournisseurs.py '
        'pour illustrer la realisation sans exposer les prix d achat ni les fournisseurs '
        'reels d un client.'
    )

    ws = wb.create_sheet(title=CATEGORISED_SHEET[:31])
    write_header(ws)
    row = 2
    for category, products in CATEGORISED_PRODUCTS.items():
        write_category_row(ws, row, category)
        row += 1
        for prod in products:
            write_product_row(ws, row, prod)
            row += 1

    for sheet_name, products in SIMPLE_SHEETS.items():
        ws = wb.create_sheet(title=sheet_name[:31])
        write_header(ws)
        for offset, prod in enumerate(products):
            write_product_row(ws, 2 + offset, prod)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument('--out', type=Path, default=DEFAULT_OUT)
    args = parser.parse_args()
    path = build(args.out)
    sheets = 1 + len(SIMPLE_SHEETS)
    rows = sum(len(p) for p in CATEGORISED_PRODUCTS.values()) + sum(
        len(p) for p in SIMPLE_SHEETS.values()
    )
    print(f'{path} — {sheets} onglets, {rows} lignes produit (donnees fictives)')


if __name__ == '__main__':
    main()
